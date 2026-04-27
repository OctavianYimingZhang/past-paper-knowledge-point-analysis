"""Render analysis outputs to Excel, JSON, Markdown, and Word.

The report-writer package is split across:

- ``__init__.py`` — the public API plus the Excel/JSON writers (which were
  unchanged in the new layout).
- ``_docx.py`` — the DOCX writer with executive summary, per-KP cheat-sheets,
  sensitivity, and appendices.
- ``_markdown.py`` — the Markdown writer that mirrors the same section order.
- ``_common.py`` — shared constants, helpers, summary derivations.

Design constraints:

- Every user-facing artifact must expose the hyperparameters used.
- The DOCX exporter is the canonical revision-plan deliverable. It is laid
  out as a one-page executive summary up front, per-KP cheat-sheets in the
  body, sensitivity warnings in a compact section, and full numeric audit
  tables plus methodology in the appendices at the end.
- Markdown mirrors the same section order so the two formats stay in sync.
- The word "conjugate" never appears in output; the model is a
  moment-matched Beta posterior. The pattern layer never claims a credible
  interval — its wording is "frequency + saturation + freshness".
- Default report language is English. Bilingual or Chinese-only output is
  opt-in via the ``lang`` parameter (``"en"``, ``"zh"``, or ``"both"``).
"""
from __future__ import annotations

import json
from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..sensitivity import (
    LeaveOneOutResult,
    SensitivitySweep,
    summarize_loo_for_report,
    summarize_sweep_for_report,
)
from ..statistical_model import KPPosterior
from ._common import (
    APPENDIX_AUDIT,
    APPENDIX_METHODOLOGY,
    APPENDIX_PATTERNS,
    SECTION_CHEAT_SHEETS,
    SECTION_EXEC_SUMMARY,
    SECTION_PREDICTIONS,
    SECTION_SENSITIVITY,
    ReportLang,
)
from ._docx import write_docx
from ._markdown import write_markdown


HEADER_FILL = PatternFill(
    start_color="FF305496", end_color="FF305496", fill_type="solid"
)
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _default(obj):
    if is_dataclass(obj):
        return asdict(obj)
    if isinstance(obj, set):
        return sorted(obj)
    if isinstance(obj, Path):
        return str(obj)
    raise TypeError(f"Unserializable type {type(obj).__name__}")


def write_json(
    out_path: str | Path,
    posteriors: list[KPPosterior],
    sweeps: dict[str, SensitivitySweep],
    loo: dict[str, LeaveOneOutResult],
    hyperparameters: dict[str, object],
) -> Path:
    """Write a single JSON payload with everything needed to reproduce the run."""
    path = Path(out_path)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "hyperparameters": hyperparameters,
        "posteriors": [asdict(p) for p in posteriors],
        "sensitivity_sweeps": {kp: asdict(s) for kp, s in sweeps.items()},
        "leave_one_out": {
            kp: {
                "baseline": asdict(r.baseline),
                "per_year": [[year, asdict(pos)] for year, pos in r.per_year],
                "max_abs_shift": r.max_abs_shift,
                "tier_flips": list(r.tier_flips),
            }
            for kp, r in loo.items()
        },
    }
    path.write_text(json.dumps(payload, indent=2, default=_default))
    return path


def write_excel(
    out_path: str | Path,
    posteriors: list[KPPosterior],
    sweeps: dict[str, SensitivitySweep],
    loo: dict[str, LeaveOneOutResult],
    hyperparameters: dict[str, object],
) -> Path:
    """Write the canonical Excel workbook."""
    path = Path(out_path)
    wb = Workbook()
    wb.remove(wb.active)

    _write_method_sheet(wb, hyperparameters)
    _write_predictions_sheet(wb, posteriors)
    _write_sensitivity_sheet(wb, sweeps)
    _write_loo_sheet(wb, loo)
    _write_trend_sheet(wb, posteriors)
    _write_review_sheet(wb, posteriors)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Excel sheet builders
# ---------------------------------------------------------------------------


def _write_method_sheet(wb: Workbook, hyperparameters: dict[str, object]) -> None:
    ws = wb.create_sheet("Method")
    rows: list[list[object]] = [
        ["Generated at", datetime.now().isoformat(timespec="seconds")],
        ["Posterior model", "Moment-matched Beta (not strict conjugate)"],
        ["Recency weighting", "w_i = exp(-lambda * (reference_year - year_i))"],
        ["Prior", "Beta(tau * coverage, tau * (1 - coverage)), tau in [0, 2]"],
        ["Trend", "Split-halves bootstrap of rate difference (no Mann-Kendall)"],
        ["Tier rules", "See references/tier-definitions.md"],
    ]
    for key, value in sorted(hyperparameters.items()):
        rows.append([key, _scalar(value)])
    _write_sheet(ws, ["Field", "Value"], rows)


def _scalar(value: object) -> object:
    """Coerce lists and other non-primitive values into an Excel-safe scalar."""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)
    return value


def _write_predictions_sheet(wb: Workbook, posteriors: list[KPPosterior]) -> None:
    ws = wb.create_sheet("Posterior_Predictions")
    headers = [
        "kp_id",
        "n_papers",
        "raw_hits",
        "weighted_hits",
        "weighted_N",
        "lambda_used",
        "tau_used",
        "coverage_share",
        "prior_alpha",
        "prior_beta",
        "posterior_alpha",
        "posterior_beta",
        "posterior_mean",
        "ci_lower_95",
        "ci_upper_95",
        "hotness_mean_share",
        "hotness_std_share",
        "trend_label",
        "trend_delta",
        "trend_ci_low",
        "trend_ci_high",
        "tier",
        "tier_reasons",
        "sensitivity_band",
        "warnings",
    ]
    rows = [
        [
            p.kp_id,
            p.n_papers,
            p.raw_hits,
            round(p.weighted_hits, 4),
            round(p.weighted_N, 4),
            p.lambda_used,
            p.tau_used,
            round(p.coverage_share, 4),
            round(p.prior_alpha, 4),
            round(p.prior_beta, 4),
            round(p.posterior_alpha, 4),
            round(p.posterior_beta, 4),
            round(p.posterior_mean, 4),
            round(p.ci_lower_95, 4),
            round(p.ci_upper_95, 4),
            round(p.hotness_mean_share, 4),
            round(p.hotness_std_share, 4),
            p.trend_label,
            round(p.trend_delta, 4),
            round(p.trend_ci_95[0], 4),
            round(p.trend_ci_95[1], 4),
            p.tier,
            " | ".join(p.tier_reasons),
            p.sensitivity_band,
            " | ".join(p.warnings),
        ]
        for p in posteriors
    ]
    _write_sheet(ws, headers, rows)


def _write_sensitivity_sheet(wb: Workbook, sweeps: dict[str, SensitivitySweep]) -> None:
    ws = wb.create_sheet("Sensitivity_Sweep")
    headers = [
        "kp_id",
        "lambda",
        "tau",
        "posterior_mean",
        "ci_lower_95",
        "ci_upper_95",
        "tier",
        "warnings",
        "band",
    ]
    rows: list[list[object]] = []
    for sweep in sweeps.values():
        for r in summarize_sweep_for_report(sweep):
            rows.append([r[h] for h in headers])
    _write_sheet(ws, headers, rows)


def _write_loo_sheet(wb: Workbook, loo: dict[str, LeaveOneOutResult]) -> None:
    ws = wb.create_sheet("Leave_One_Out")
    headers = [
        "kp_id",
        "dropped_year",
        "baseline_posterior_mean",
        "loo_posterior_mean",
        "shift",
        "abs_shift",
        "baseline_tier",
        "loo_tier",
        "tier_flipped",
    ]
    rows: list[list[object]] = []
    for result in loo.values():
        for r in summarize_loo_for_report(result):
            rows.append([r[h] for h in headers])
    _write_sheet(ws, headers, rows)


def _write_trend_sheet(wb: Workbook, posteriors: list[KPPosterior]) -> None:
    ws = wb.create_sheet("Trend_Analysis")
    headers = [
        "kp_id",
        "trend_label",
        "trend_delta",
        "trend_ci_low",
        "trend_ci_high",
        "historical_mean",
        "posterior_mean",
    ]
    rows = [
        [
            p.kp_id,
            p.trend_label,
            round(p.trend_delta, 4),
            round(p.trend_ci_95[0], 4),
            round(p.trend_ci_95[1], 4),
            round(p.historical_mean, 4),
            round(p.posterior_mean, 4),
        ]
        for p in posteriors
    ]
    _write_sheet(ws, headers, rows)


def _write_review_sheet(wb: Workbook, posteriors: list[KPPosterior]) -> None:
    ws = wb.create_sheet("Review_Queue")
    headers = ["kp_id", "tier", "warning"]
    rows: list[list[object]] = []
    for p in posteriors:
        for w in p.warnings:
            rows.append([p.kp_id, p.tier, w])
    _write_sheet(ws, headers, rows)


def _write_sheet(ws, headers: list[str], rows: Iterable[list[object]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for row in rows:
        ws.append(list(row))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
    for column_cells in ws.columns:
        width = 14
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, 60))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


__all__ = (
    "APPENDIX_AUDIT",
    "APPENDIX_METHODOLOGY",
    "APPENDIX_PATTERNS",
    "ReportLang",
    "SECTION_CHEAT_SHEETS",
    "SECTION_EXEC_SUMMARY",
    "SECTION_PREDICTIONS",
    "SECTION_SENSITIVITY",
    "write_docx",
    "write_excel",
    "write_json",
    "write_markdown",
)
